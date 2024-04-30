from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook('pivot_table.xlsx')
pt_sheet = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()
data = Reference(pt_sheet,
                 min_col=min_col + 1,
                 max_col=max_col,
                 min_row=min_row,
                 max_row=max_row)

categories = Reference(pt_sheet,
                       min_col=min_col,
                       max_col=min_col,
                       min_row=min_row + 1,
                       max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

pt_sheet.add_chart(barchart, 'A7')

barchart.title = "Sales by Product Line"
barchart.style = 5
wb.save('barchart.xlsx')
