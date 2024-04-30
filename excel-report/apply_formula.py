from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('barchart.xlsx')
pt_sheet = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# formula = '=SUM(B3:B4)'
# pt_sheet['B5'] = formula
# pt_sheet['B5'].style = 'Currency'

for i in range(min_col+1, max_col+1):
    letter = get_column_letter(i)
    pt_sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    pt_sheet[f'{letter}{max_row+1}'].style = 'Currency'
wb.save('report.xlsx')
